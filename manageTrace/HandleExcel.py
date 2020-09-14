#!/usr/bin/python
# -*- coding: UTF-8 -*-


from openpyxl import load_workbook
from openpyxl import Workbook
import chardet
import os, time
import json
import requests
from testCreateSrv.Logger import Logger

url = "https://qianbaidugg.utools.club/modify/"
url = 'http://127.0.0.1:8000/modify/'
file_path = u'D:/userdata/WeChat Files/WeChat Files/wxid_f56prkbc2zpm21/FileStorage/File/2020-07/任务数据详细/'

param = {"personnel_id": "1",
         "function": "addCompleteItemMsg",
         "ItemList": [] }

log_debug = Logger('import_excel.log', level='debug')
def write_debug(s):
    log_debug.logger.debug(s)


def read_file_list():
    names = os.listdir(file_path)
    l = []
    for i in names:
        if i.find('.xlsx'):
            l.append(i)
        else:
            continue
    return l

# "area_name": "河北",
# "product_name": "维保",
# "personnel_name": "吴文浩",
# "product_type": "主设备",
# "item_name": "武汉电信PON备件服务项目",
# "item_money": 98.00,
# "item_money_fact": 100.00,
# "item_type": 0,
# "item_state": 0,
# "item_frame": 0,
# "item_record_time": "2020-07-15 9:03:13",
# "item_modify_time": "2020-07-20 9:03:13"

def send_data(body):
    headers = {'Content-Type': 'application/json'}
    write_debug(body)
    datas = json.dumps(body)
    r = requests.post(url, data=datas, headers=headers)
    write_debug(r.text)
    

main_p = [u'维保', u'网优', u'备件', u'哑资源',u'培训']
def deal_data(data):
    for d in data:
        key_vars = {}
        tm = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
        for v in d:
            if v.find(u'省份') >= 0 or v.find(u'区域') >= 0:
                key_vars['area_name'] = d[v]
            elif v.find(u'产品') >= 0 :
                key_vars['product_name'] = d[v]
                if d[v] in main_p:
                    key_vars['product_type'] = u'主设备'
                else:
                    key_vars['product_type'] = u'软件类'
            elif v.find(u'项目名称') >= 0 :
                key_vars['item_name'] = d[v] 
            elif v.find(u'是否框架') >= 0 :
                if d[v] == u'是':
                    key_vars['item_frame'] = 0
                else:
                    key_vars['item_frame'] = 1
            elif v.find(u'类型') >= 0 :
                if d[v] == u'存量':
                    key_vars['item_type'] = 0
                else:
                    key_vars['item_type'] = 1
            elif v.find(u'责任人') >= 0 :
                key_vars['personnel_name'] = d[v]
            elif v.find(u'预计落地金额') >= 0 :
                key_vars['item_money'] = float(d[v])
            elif v.find(u'当年框架金额') >= 0 :
                key_vars['item_money_fact'] = 0 # float(d[v]) # 年度线索项目，录入的时候实际金额为0
            else:
                continue
        #end
        key_vars['item_record_time'] = tm
        key_vars['item_modify_time'] = tm
        key_vars['item_state'] = 0 #导入的项目全为未完成
        key_vars['item_flag'] = 0  # Excel导入的全是年度线索
        param['ItemList'].append(key_vars)
    #end    
    send_data(param)
    
    return

def deal_sheet(wb, st):
    sheet = wb[st]
    write_debug(sheet.title + ':/n')
    it_names = []
    sheet_data = []
    for r in range(sheet.min_row, sheet.max_row + 1):
        key_vars = {}
        for c in range(sheet.min_column, sheet.max_column + 1):
            if r == sheet.min_row :
                it_names.append(str(sheet.cell(row=r, column=c).value))
            else:
                key_vars[it_names[c-1]] = str(sheet.cell(row=r, column=c).value)

        if r != sheet.min_row :
            sheet_data.append(key_vars)
            
    #end 
    deal_data(sheet_data)
    return

def do_file():
    # files = read_file_list()
    # write_debug(files)
    #
    # for i in files:
    file = 'E:/src_code/gg_test/testCreateSrv/' + '丁洋-项目路径表.xlsx'
    write_debug (file)
    wb = load_workbook(file)

    sheets = wb.sheetnames
    write_debug(sheets)
    # 循环遍历所有sheet
    for j in range(len(sheets)):
        deal_sheet(wb, sheets[j])
    
    return

do_file()
